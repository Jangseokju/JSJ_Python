import os
import openpyxl as px
from datetime import date
import re

# 파일 경로 입력
program_path = os.path.abspath(__file__)
path = os.path.dirname(program_path)
input_file = path + '/input/' + input('입력 파일 : ')
output_file = path + '/output/' + input('출력 파일 : ')

# 엑셀 파일 입력
workbook = px.load_workbook(input_file, read_only=True)
worksheet = workbook['january_2013']

# 엑셀 출력 객체 생성
output_workbook = px.Workbook()
output_worksheet = output_workbook.active
output_workbook.title = 'out_january_2013'

# 정규표현식으로 필터링
# * Customer Name 이 대문자 J 로 시작하는 행만 필터링
pattern = re.compile(r'(?P<my_pattern>^J.*)')

index = 1

workbook = px.load_workbook(input_file, read_only=True)
worksheet = workbook['january_2013']

data = []
header = [cell.value for cell in worksheet[1]]
data.append(header)

for row in worksheet.iter_rows(min_row=2, values_only=True):
    row_list = []
    # 정규표현식 패턴에 일치하는지 확인
    # row[index] : Customer Name
    if pattern.match( row[index] ):
        for cell_value in row:
            row_list.append(cell_value)
    if row_list:
        data.append(row_list)


# 엑셀 파일 출력
for row_index, output_list in enumerate(data, 1):
    for col_index, value in enumerate(output_list, 1):
        output_worksheet.cell(row=row_index, column=col_index, value=value)

output_workbook.save(output_file)

