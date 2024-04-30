from openpyxl import load_workbook      # 엑셀 입력
from openpyxl import Workbook           # 엑셀 출력         💛 추가
import os

# 현재 실행 파일 경로 가져오고, 입력파일 지정하기
program_path = os.path.abspath(__file__)
path = os.path.dirname(program_path)
input_file = path + '/input/' + input('입력 파일 : ')
output_file = path + '/output/' + input('출력 파일 : ')       # 💛 추가


# 엑셀 통합 문서 열기 (입력)
workbook = load_workbook(input_file)
# january_2013 워크시트만 입력
worksheet = workbook['january_2013']

# 엑셀 출력 객체 생성
output_workbook = Workbook()
output_worksheet = output_workbook.active
output_worksheet.title = ' out_january_2013 '

# sales_2013.xlsx 의 january_2013 워크시트를 반복하여
# output02.xlsx 의 out_january_2013 워크시트로 출력
# 행 반복
for row_index, row in enumerate( worksheet.iter_rows(), 1 ):
    # 열 반복
    for column_index, cell in enumerate(row, 1):
        # output_worksheet.cell(row=행, column=열, value=값)
        output_worksheet.cell(row=row_index, column=column_index, value=cell.value)


# 엑셀 통합 문서 저장
output_workbook.save(output_file)