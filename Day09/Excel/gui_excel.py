
from openpyxl import load_workbook      # 엑셀 입력
from openpyxl import Workbook           # 엑셀 출력      

import tkinter as tk
from tkinter import filedialog

# 입력 파일 선택 버튼 클릭 시,
def open_input_file():
    print('입력 파일 선택...')
    # 파일 선택 상자에서 파일이름 가져오기
    filename = filedialog.askopenfilename()
    if filename:
        # 파일명 지정
        input_file.set(filename)


# 출력 파일 선택 버튼 클릭 시,
def open_output_file():
    print('출력 파일 선택...')
    # 파일 선택 상자에서 파일이름 가져오기
    filename = filedialog.asksaveasfilename()
    if filename:
        # 파일명 지정
        output_file.set(filename)

# 실행 버튼
def run():
    print('입력 파일 경로 : ' + input_file.get())
    print('출력 파일 경로 : ' + output_file.get())
    print('데이터 분석을 시작합니다...')
    work()

# 데이터 분석
def work():
    # 엑셀 통합 문서 열기 (입력)
    workbook = load_workbook(input_file.get())
    # january_2013 워크시트만 입력
    worksheet = workbook['january_2013']

    # 엑셀 출력 객체 생성
    output_workbook = Workbook()
    output_worksheet =  output_workbook.active  # 워크시트 활성화
    output_worksheet.title = 'out_january_2013' # 워크시트 이름 지정

    # ✅ TODO :분석
    for row_index, row in  enumerate( worksheet.iter_rows(), 1 ):
        # 열 반복
        for column_index, cell in enumerate(row, 1 ):
            output_worksheet.cell(row=row_index, column=column_index, value=cell.value)


    # 엑셀 통합 문서 저장
    output_workbook.save(output_file.get())




# 윈도우 화면 생성
window = tk.Tk()
window.title('GUI 프로그램')

# 창 크기 지정
window.geometry('600x300')

# 라벨 생성
label = tk.Label(window, text="GUI 프로그램 라벨", padx=20, pady=10)
label.pack()

# 입력 상자 생성 - 입력 파일 경로
input_file = tk.StringVar()         # 문자열 변수 
input_entry = tk.Entry(window, textvariable=input_file, width=100)
input_entry.pack()

# 버튼 - 입력 파일 선택 버튼
input_button = tk.Button(window, text="입력 파일 선택", command=open_input_file)
input_button.pack()


# 입력 상자 생성 - 출력 파일 경로
output_file = tk.StringVar()         # 문자열 변수 
output_entry = tk.Entry(window, textvariable=output_file, width=100)
output_entry.pack()

# 버튼 - 입력 파일 선택 버튼
output_button = tk.Button(window, text="출력 파일 선택", command=open_output_file)
output_button.pack()

# 실행 버튼
run_button = tk.Button(window, text="실행", padx=10, pady=10, command=run)
run_button.pack()

window.mainloop()