import os
import openpyxl as px
from datetime import date
import re

# 파일 경로 입력
program_path = os.path.abspath(__file__)
path = os.path.dirname(program_path)
input_file = path + '/input/' + input('입력 파일 : ')
output_file = path + '/output/' + input('출력 파일 : ')

# 엑셀 파일 입력
workbook = px.load_workbook(input_file, read_only=True)
worksheet = workbook['january_2013']

# 엑셀 출력 객체 생성
output_workbook = px.Workbook()
output_worksheet = output_workbook.active
output_workbook.title = 'out_january_2013'

# Customer Name, Sales Amount
select_index = [1, 3]

data = []
for row in worksheet.iter_rows(values_only=True):
    row_list = []
    for column_index in select_index:
        cell_value = row[column_index]
        row_list.append(cell_value)
    data.append(row_list)


