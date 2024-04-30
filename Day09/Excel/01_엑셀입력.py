from openpyxl import load_workbook
from openpyxl import Workbook
import os

# 현재 실행 파일 경로 가져오고, 입력파일 지정하기
path = 'C:/ALOHA/SEOKJU/GIT/JSJ_Python/Day09/Excel/'
input_file = path + '/input/' + input('입력 파일 : ')
# sales_2013.xlsx

# 엑셀 파일 로드
# load_workbook : 지정한 엑셀 파일 읽어와서 WorkBook 객체로 가져온다.
workbook = load_workbook(input_file)
print('엑셀 통합 문서의 워크시트 수 : ', len(workbook.sheetnames))
for worksheet in workbook:
    print('워크시트 이름 : ', worksheet.title)
    print('워크시트 최대 행 수 : ', worksheet.max_row)
    print('워크시트 최대 열 수 : ', worksheet.max_column)

# 첫 번째 시트 선택
sheet = workbook.active

# 시트 내용 출력
# iter_rows() : 시트에서 행들을 반복가능한 객체로 반환
for row in sheet.iter_rows(values_only=True):
    print(row)