import os
import openpyxl as px
from datetime import datetime

# 파일 경로 입력
program_path = os.path.abspath(__file__)
path = os.path.dirname(program_path)
input_file = path + '/input/' + input('입력 파일 : ')
output_file = path + '/output/' + input('출력 파일 : ')

# 엑셀 파일 입력
workbook = px.load_workbook(input_file, read_only=True)
worksheet = workbook['january_2013']

# 구매일자 - purchase_date : 특정 날짜들만(집합)
# 집합으로 필터링
important_dates = ['01/06/2013', '01/11/2013']

index = 4       # purchase_date 열의 인덱스

# 엑셀 출력 객체 생성
output_workbook = px.Workbook()
output_worksheet = output_workbook.active
output_workbook.title = 'out_january_2013'

# 출력할 데이터 리스트
data = []
header = [cell.value for cell in worksheet[1]]
data.append(header)

# 필터링
# - 1행은 헤더, 2행부터 데이터를 가져와서 반복
for row in worksheet.iter_rows(min_row=2, values_only=True):
    purchase_date = row[index]
    purchase_date = purchase_date.strftime('%m/%d/%Y')
    # 미리 지정한 집합만 선택
    if purchase_date in important_dates:
        # 출력 데이터 리스트에 추가
        data.append(list(row))

# 출력 데이터를 엑셀 파일로 저장
for row_index, row in  enumerate( data, 1 ):
    for column_index, value in enumerate(row, 1 ):
        if type(value) == datetime:
            value = value.strftime('%Y/%m/%d')
            output_worksheet.cell(row=row_index, column=column_index, value=value)
        else:
            output_worksheet.cell(row=row_index, column=column_index, value=value)

# 엑셀 통합 문서 저장
output_workbook.save(output_file)




