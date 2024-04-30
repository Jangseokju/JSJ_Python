from openpyxl import load_workbook      # 엑셀 입력
from openpyxl import Workbook           # 엑셀 출력         
import os

# 현재 실행 파일 경로 가져오고, 입력파일 지정하기
program_path = os.path.abspath(__file__)
path = os.path.dirname(program_path)
input_file = path + '/input/' + input('입력 파일 : ')
output_file = path + '/output/' + input('출력 파일 : ')      


# 엑셀 통합 문서 열기 (입력)
workbook = load_workbook(input_file)
worksheet_name = input('워크시트 이름 : ')
worksheet = workbook[worksheet_name]

# 엑셀 출력 객체 생성
output_workbook = Workbook()
output_worksheet =  output_workbook.active  # 워크시트 활성화
output_worksheet.title = input('출력 워크시트 이름 : ')

# 출력 데이터
data = []
# 리스트 내포 - 1행의 모든 셀을 반복하여 리스트로 반환
header = [cell.value for cell in worksheet[1]]

# 헤더 추가
data.append(header)

# 필터링
# - 1행은 헤더, 2행부터 데이터를 가져와서 반복
for row in worksheet.iter_rows(min_row=2, values_only=True):
    # row[3] : Sale Amount 열 데이터를 가져와서 1500 보다 크면
    if float( row[3] ) > 1500.0:
        # 출력 데이터 리스트에 추가
        data.append(list(row))

# 출력 데이터를 엑셀 파일로 저장
for row_index, row in  enumerate( data, 1 ):
    for column_index, value in enumerate(row, 1 ):
        output_worksheet.cell(row=row_index, column=column_index, value=value)

# 엑셀 통합 문서 저장
output_workbook.save(output_file)

