import os
import openpyxl as px
from datetime import date
import re

# 파일 경로 입력
program_path = os.path.abspath(__file__)
path = os.path.dirname(program_path)
input_file = path + '/input/' + input('입력 파일 : ')
output_file = path + '/output/' + input('출력 파일 : ')

# 엑셀 파일 입력
workbook = px.load_workbook(input_file, read_only=True)
worksheet = workbook['january_2013']

# 엑셀 출력 객체 생성
output_workbook = px.Workbook()
output_worksheet = output_workbook.active
output_workbook.title = 'out_january_2013'

# 선택할 헤더
select_header = ['Customer Name', 'Sale Amount']

header_list = worksheet[1]
header_index_list = []

for header_index in range(len(header_list)):
    # 헤더를 반복하여, 선택한 헤더명과 일치하는 index 만 추출
    if header_list[header_index].value in select_header:
        header_index_list.append(header_index)

data = []
for row in worksheet.iter_rows(values_only=True):
    row_list = []
    for column_index in header_index_list:
        cell_value = row[column_index]
        row_list.append(cell_value)
    data.append(row_list)

# 엑셀 파일 출력
for row_index, output_list in enumerate(data, 1):
    for col_index, value in enumerate(output_list, 1):
        output_worksheet.cell(row=row_index, column=col_index, value=value)

output_workbook.save(output_file)




